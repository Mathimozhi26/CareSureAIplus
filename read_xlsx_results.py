from pathlib import Path
from openpyxl import load_workbook
path = Path('Vulnerability Test Results') / 'Skin_Journey_QA_Test_Cases.xlsx'
wb = load_workbook(path)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
idx = headers.index('Pass/Fail') if 'Pass/Fail' in headers else None
count = 0
passed = 0
failed = 0
blank = 0
other = []
for row in rows[1:]:
    count += 1
    val = row[idx] if idx is not None else ''
    text = str(val).strip() if val is not None else ''
    if text.lower() == 'pass':
        passed += 1
    elif text.lower() == 'fail':
        failed += 1
    elif text == '':
        blank += 1
    else:
        other.append((count, text))
print(count, passed, failed, blank, len(other))
for item in other[:20]:
    print(item)
